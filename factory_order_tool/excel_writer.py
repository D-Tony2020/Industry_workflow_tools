"""Excel输出模块 - 按「导入产品明细模板」格式输出xlsx"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from config import OUTPUT_COLUMNS


def write_output_excel(output_rows, output_path):
    """
    将映射后的订单数据按工厂系统模板格式写入Excel。

    参数:
        output_rows: list[dict] - 包含OUTPUT_COLUMNS字段的行列表
        output_path: str - 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 样式
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头（与模板完全一致）
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 写入数据行
    for row_idx, row_data in enumerate(output_rows, 2):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            value = row_data.get(col_name, "")

            # 数字字段转换
            if col_name in ("数量", "采购单价") and value:
                try:
                    value = float(str(value).replace(",", ""))
                except (ValueError, TypeError):
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 设置列宽
    col_widths = {
        "产品编号": 14, "产品名称": 35, "产品规格": 14, "数量": 10,
        "计划开始时间": 14, "计划结束时间": 14, "工艺路线名称": 18,
        "工序列表": 10, "备注": 30, "更新": 6, "工单分类": 10,
        "供应商": 10, "供应商名称": 15, "供应商联系人": 12,
        "供应商联系电话": 14, "收货地址": 15, "采购单价": 10,
        "客户选择": 10, "关联产品": 10,
    }
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
        letter = _col_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 12)

    wb.save(output_path)
    wb.close()


def _col_letter(col_idx):
    """将列号转为Excel列字母（1=A, 27=AA）"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result
