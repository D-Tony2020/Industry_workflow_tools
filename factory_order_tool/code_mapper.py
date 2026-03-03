"""编码映射模块 - 读取料号清单Excel，执行客户编码到工厂编码的映射"""
import os
from datetime import datetime, date
from openpyxl import load_workbook
from config import (
    MAPPING_TABLE_PATH,
    MAP_COL_IDX_JY_CODE,
    MAP_COL_IDX_CUSTOMER,
    MAP_COL_IDX_DESC,
    OUTPUT_ORDER_TYPE,
    QUANTITY_SAFETY_MARGIN,
)


def load_mapping_table(path=None):
    """
    读取料号清单Excel（多sheet），返回映射字典。

    新结构（每个sheet）：
      Row 1: 标题或空行
      Row 2: 表头（序号 | 久益料号 | <客户>料号 | 品名规格 | 备注）
      Row 3+: 数据

    映射键: 客户料号（如 YY60030058）
    映射值: dict 包含产品编号(久益料号)、产品名称(品名规格)

    返回:
        mapping: dict - {客户料号: {产品编号, 产品名称}}
    """
    path = path or MAPPING_TABLE_PATH
    if not os.path.exists(path):
        return {}

    wb = load_workbook(path, read_only=True)
    mapping = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_found = False

        for row in ws.iter_rows(values_only=True):
            # 查找表头行（含"久益料号"的行）
            if not header_row_found:
                cells = [str(c).strip() if c else "" for c in row]
                if any("久益料号" in c for c in cells):
                    header_row_found = True
                continue

            # 数据行
            if row is None or len(row) <= MAP_COL_IDX_DESC:
                continue

            raw_jy = row[MAP_COL_IDX_JY_CODE]
            raw_customer = row[MAP_COL_IDX_CUSTOMER]
            raw_desc = row[MAP_COL_IDX_DESC]

            # 转字符串，处理数值型客户料号（如 sheet 0005 中的整数）
            jy_code = _to_str(raw_jy)
            customer_code = _to_str(raw_customer)
            desc = _to_str(raw_desc)

            if not jy_code or not customer_code:
                continue

            mapping[customer_code] = {
                "产品编号": jy_code,
                "产品名称": desc,
            }

    wb.close()
    return mapping


def _to_str(value):
    """将单元格值安全转为字符串，处理 None / int / float"""
    if value is None:
        return ""
    if isinstance(value, float):
        # 避免 2174490103.0 → "2174490103.0"
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def apply_mapping(items, mapping):
    """
    对解析出的订单项目应用编码映射，生成输出行。

    v1.1.0 输出规则：
    - 仅填充5个必填字段：产品编号、产品规格、数量(+5)、计划开始时间、工单分类
    - 其余14列留空

    参数:
        items: list[dict] - PDF解析出的项目列表
        mapping: dict - 映射字典

    返回:
        output_rows: list[dict] - 输出模板格式的行列表
        unmapped: list[str] - 未找到映射的料件编号列表
    """
    output_rows = []
    unmapped = []

    today_str = date.today().strftime("%Y/%m/%d")

    for item in items:
        customer_code = item.get("料件编号", "").strip()
        product_info = mapping.get(customer_code)

        # 数量 + 安全余量
        raw_qty = item.get("采购数量", "")
        try:
            qty = int(float(str(raw_qty).replace(",", ""))) + QUANTITY_SAFETY_MARGIN
        except (ValueError, TypeError):
            qty = raw_qty  # 无法转换时保留原值

        # 计划结束时间 = 订单交货日期（若早于今天则用今天）
        delivery_date_str = item.get("出货日期", "").strip()
        end_date = _resolve_end_date(delivery_date_str, today_str)

        row = {
            # ===== 5个必填字段 =====
            "产品编号": "",
            "产品规格": customer_code,
            "数量": qty,
            "计划开始时间": today_str,
            "工单分类": OUTPUT_ORDER_TYPE,
            # ===== 其余列留空 =====
            "产品名称": "",
            "计划结束时间": end_date,
            "工艺路线名称": "",
            "工序列表": "",
            "备注": "",
            "更新": "",
            "供应商": "",
            "供应商名称": "",
            "供应商联系人": "",
            "供应商联系电话": "",
            "收货地址": "",
            "采购单价": "",
            "客户选择": "",
            "关联产品": "",
            # ===== 内部字段（预览/比对用，不写入导出Excel）=====
            "_映射状态": "未映射",
            "_产品名称": "",
            "_品名": item.get("品名", ""),
            "_图号": item.get("图号", ""),
            "_规格": item.get("规格", ""),
            "_交期回复": item.get("交期回复", ""),
            "_项次": item.get("项次", ""),
        }

        if product_info:
            row["产品编号"] = product_info["产品编号"]
            row["_产品名称"] = product_info["产品名称"]
            row["_映射状态"] = "已映射"
        else:
            if customer_code:
                unmapped.append(customer_code)

        output_rows.append(row)

    return output_rows, unmapped


def _resolve_end_date(delivery_date_str, today_str):
    """
    解析交货日期，若早于今天则返回今天。

    支持格式: "2026/03/28", "2026-03-28" 等常见日期格式

    参数:
        delivery_date_str: str - 订单中的交货日期
        today_str: str - 今天日期 (YYYY/MM/DD)

    返回:
        str - 有效的结束日期 (YYYY/MM/DD)
    """
    if not delivery_date_str:
        return today_str

    # 尝试解析交货日期
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            delivery_date = datetime.strptime(delivery_date_str, fmt).date()
            today = datetime.strptime(today_str, "%Y/%m/%d").date()
            if delivery_date < today:
                return today_str
            return delivery_date.strftime("%Y/%m/%d")
        except ValueError:
            continue

    # 无法解析时原样返回
    return delivery_date_str


def get_mapping_stats(output_rows):
    """统计映射结果"""
    total = len(output_rows)
    mapped = sum(1 for r in output_rows if r.get("_映射状态") == "已映射")
    return total, mapped, total - mapped
